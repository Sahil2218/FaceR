import cv2
import subprocess
import time
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook

# Load the pre-trained Haar Cascade face detector
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize the webcam
cap = cv2.VideoCapture(1)

face_detected = False

# Path to the folder where you want to save the detected face images
save_folder = 'FACES'

# Create the folder if it doesn't exist
os.makedirs(save_folder, exist_ok=True)

# Ask the user to input their name
name = input("Enter your name: ")

# Load the existing Excel workbook or create a new one if it doesn't exist
try:
    workbook = load_workbook('user_data.xlsx')
    worksheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet['A1'] = 'Face'
    worksheet['B1'] = 'Email'
    worksheet['C1'] = 'Password'

while True:
    # Capture a frame
    ret, frame = cap.read()

    # Convert the frame to grayscale for face detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the frame
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

    # Draw rectangles around the detected faces
    for (x, y, w, h) in faces:
        cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)

        # Set face_detected flag if a face is detected
        if not face_detected:
            face_detected = True

            # Capture and save the image of the detected face
            face_img = frame[y:y+h, x:x+w]
            save_path = os.path.join(save_folder, f'{name}_face.jpg')
            cv2.imwrite(save_path, face_img)

            # Add face image to the Excel file
            row = ('', '', '')
            worksheet.append(row)
            img = Image(save_path)
            worksheet.add_image(img, 'A' + str(worksheet.max_row))

    # Display the live feed with detected faces
    cv2.imshow('Face Recognition', frame)

    # Break the loop if 'q' is pressed or a face is detected
    if cv2.waitKey(1) & 0xFF == ord('q') or face_detected:
        break

# Release the webcam and close the window
cap.release()
cv2.destroyAllWindows()

# Save the Excel workbook
workbook.save('user_data.xlsx')

# Open the index2.html file in a web browser after the OpenCV window is closed
if face_detected:
    subprocess.call(['open', 'index2.html'])
    time.sleep(3)  # Wait for 3 seconds